"""
itr_status_checker.pyw  —  Bulk Income-Tax Return Verification Status checker
=============================================================================
A single-file desktop tool (Tkinter GUI, no console window because of the
.pyw extension) that:

  1. Exports a blank Excel TEMPLATE  (Name | PAN | Password | AY | Status)
  2. Lets you UPLOAD the filled-in template
  3. RUNS the batch: for every row it
        logs in  ->  e-File  ->  Income Tax Returns  ->  View Filed Returns
        ->  Filter by the Assessment Year  ->  reads the LATEST status
        (e.g. "Processed with no demand/refund")  ->  records it in the
        Status column  ->  logs out  ->  next row, and repeats.
  4. Lets you DOWNLOAD an updated copy of the Excel with the Status column
     filled in (the original file is left untouched).

The Selenium login logic is the component you supplied (incometax_login.py),
embedded here so this stays a one-file deliverable.

-----------------------------------------------------------------------------
REQUIREMENTS  (install once, in a Command Prompt):
    pip install selenium openpyxl
    # Microsoft Edge must be installed (Selenium 4.6+ auto-manages the driver)

RUN:
    Double-click itr_status_checker.pyw   (or:  pythonw itr_status_checker.pyw)
-----------------------------------------------------------------------------
"""

import os
import time
import threading
import queue
import re
import json
from datetime import date, datetime

# --- GUI -------------------------------------------------------------------
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

# --- Excel -----------------------------------------------------------------
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _OPENPYXL_OK = True
except Exception:                                            # pragma: no cover
    _OPENPYXL_OK = False

# --- Selenium --------------------------------------------------------------
from selenium import webdriver
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException


# =========================================================================== #
#  PART 1 — LOGIN COMPONENT  (from your incometax_login.py, lightly trimmed)   #
# =========================================================================== #
class LoginError(Exception):
    """Login could not be completed (e.g. Continue never proceeded)."""


class InvalidPasswordError(LoginError):
    """The portal reported an invalid/incorrect password. Do NOT retry."""


class IncomeTaxPortal:
    PORTAL_URL   = "https://www.incometax.gov.in/iec/foportal/"
    LOGIN_URL    = "https://eportal.incometax.gov.in/iec/foservices/#/login"
    PAGE_TIMEOUT = 20

    def __init__(self, driver=None, headless=False, log=None):
        self.log = log or (lambda m: print(m))
        self.driver = driver or self.build_driver(headless)

    # ---- driver construction --------------------------------------------- #
    @staticmethod
    def build_driver(headless=False):
        opts = EdgeOptions()
        opts.page_load_strategy = "none"   # never block on page loads; we poll
        if headless:
            opts.add_argument("--headless=new")
        opts.add_argument("--start-maximized")
        opts.add_argument("--disable-blink-features=AutomationControlled")
        return webdriver.Edge(service=EdgeService(), options=opts)

    # ---- generic robust helpers ------------------------------------------ #
    def find(self, locators, timeout=None, need_clickable=False):
        timeout = self.PAGE_TIMEOUT if timeout is None else timeout
        end = time.time() + timeout
        while True:
            for by, val in locators:
                try:
                    els = self.driver.find_elements(by, val)
                except Exception:
                    els = []
                for el in els:
                    try:
                        if not el.is_displayed():
                            continue
                        if need_clickable and not el.is_enabled():
                            continue
                        return el
                    except Exception:
                        continue
            if time.time() >= end:
                raise TimeoutException(f"No locator matched in {timeout}s: {locators}")
            time.sleep(0.3)

    def click(self, locators, timeout=None):
        el = self.find(locators, timeout=timeout, need_clickable=True)
        try:
            el.click()
        except Exception:
            self.driver.execute_script("arguments[0].click();", el)
        return el

    def type(self, locators, text, timeout=None):
        """Type into a field fast, then verify it landed; if not, force the
        value via JS + input/change events so Angular registers it."""
        el = self.find(locators, timeout=timeout)
        try:
            self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
        except Exception:
            pass
        try:
            el.click()
        except Exception:
            pass
        try:
            el.clear()
            el.send_keys(text)
        except Exception:
            pass
        if (el.get_attribute("value") or "").strip() != str(text).strip():
            self.driver.execute_script(
                "arguments[0].value = arguments[1];"
                "arguments[0].dispatchEvent(new Event('input',{bubbles:true}));"
                "arguments[0].dispatchEvent(new Event('change',{bubbles:true}));"
                "arguments[0].dispatchEvent(new Event('blur',{bubbles:true}));",
                el, str(text))
        return el

    def click_text(self, text, timeout=10, pick="shortest"):
        """Click an element by visible text. pick='shortest' (default),
        'top' (smallest y) or 'bottom' (largest y)."""
        js = r"""
        const want = arguments[0].trim().toLowerCase();
        const pick = arguments[1];
        const tags = ['button','a','input','span','label','li','div','td'];
        let best = null;
        for (const tag of tags){
          for (const el of document.getElementsByTagName(tag)){
            const r = el.getBoundingClientRect();
            if (r.width === 0 || r.height === 0) continue;
            const txt = (el.innerText || el.value || '').trim().toLowerCase();
            if (!txt) continue;
            if (txt === want || txt.includes(want)){
              const cand = {el: el, len: txt.length, top: r.top};
              if (!best){ best = cand; }
              else if (pick === 'shortest' && cand.len < best.len){ best = cand; }
              else if (pick === 'top'      && cand.top < best.top){ best = cand; }
              else if (pick === 'bottom'   && cand.top > best.top){ best = cand; }
            }
          }
        }
        if (!best) return false;
        best.el.scrollIntoView({block:'center'});
        best.el.click();
        return true;
        """
        end = time.time() + timeout
        while time.time() < end:
            try:
                if self.driver.execute_script(js, text, pick):
                    return True
            except Exception:
                pass
            time.sleep(0.3)
        raise TimeoutException(f"No clickable element with text '{text}'")

    def click_css(self, css, timeout=15):
        """Click the first visible element matching a CSS selector (fast path
        for elements with a stable id/class). JS click, so overlays don't block."""
        js = r"""
        const el = document.querySelector(arguments[0]);
        if (!el) return false;
        const r = el.getBoundingClientRect();
        if (r.width === 0 || r.height === 0) return false;
        el.scrollIntoView({block:'center'});
        el.click();
        return true;
        """
        end = time.time() + timeout
        while time.time() < end:
            try:
                if self.driver.execute_script(js, css):
                    return True
            except Exception:
                pass
            time.sleep(0.2)
        raise TimeoutException(f"No clickable element for CSS '{css}'")

    def click_menu_item(self, text, timeout=12):
        """Click a Material menu item (role=menuitem / .mat-mdc-menu-item) by
        its visible text — used for the e-File submenu and the Log Out item."""
        js = r"""
        const want = arguments[0].trim().toLowerCase();
        const items = document.querySelectorAll(
            '[role=menuitem], .mat-mdc-menu-item, button[mat-menu-item]');
        for (const el of items){
          const r = el.getBoundingClientRect();
          if (r.width === 0 || r.height === 0) continue;
          const t = (el.innerText || el.textContent || '').trim().toLowerCase();
          if (t === want || t.includes(want)){
            el.scrollIntoView({block:'center'});
            el.click();
            return true;
          }
        }
        return false;
        """
        end = time.time() + timeout
        while time.time() < end:
            try:
                if self.driver.execute_script(js, text):
                    return True
            except Exception:
                pass
            time.sleep(0.2)
        raise TimeoutException(f"No menu item '{text}'")

    def check_box(self, timeout=8):
        def is_ticked():
            for b in self.driver.find_elements(By.XPATH, "//input[@type='checkbox']"):
                try:
                    if b.is_selected():
                        return True
                except Exception:
                    pass
            for b in self.driver.find_elements(
                    By.XPATH, "//*[@role='checkbox' or contains(@class,'p-checkbox-box')"
                              " or contains(@class,'mat-checkbox')]"):
                try:
                    cls = b.get_attribute("class") or ""
                    if (b.get_attribute("aria-checked") == "true"
                            or "p-highlight" in cls or "checked" in cls):
                        return True
                except Exception:
                    pass
            return False

        selectors = [
            "//label[contains(translate(.,'SECURE','secure'),'secure access')]",
            "//label[contains(translate(.,'AGREE','agree'),'agree')]",
            "//*[contains(@class,'p-checkbox-box')]",
            "//*[@role='checkbox']",
            "//*[contains(@class,'mat-checkbox-inner-container')]",
            "//label[.//input[@type='checkbox']]",
            "//input[@type='checkbox']",
        ]
        end = time.time() + timeout
        while time.time() < end:
            if is_ticked():
                return True
            for sel in selectors:
                for el in self.driver.find_elements(By.XPATH, sel):
                    try:
                        if not el.is_displayed():
                            continue
                        self.driver.execute_script(
                            "arguments[0].scrollIntoView({block:'center'});", el)
                        try:
                            el.click()
                        except Exception:
                            self.driver.execute_script("arguments[0].click();", el)
                        if is_ticked():
                            return True
                    except Exception:
                        continue
            for b in self.driver.find_elements(By.XPATH, "//input[@type='checkbox']"):
                try:
                    self.driver.execute_script(
                        "arguments[0].checked = true;"
                        "arguments[0].dispatchEvent(new Event('click',{bubbles:true}));"
                        "arguments[0].dispatchEvent(new Event('change',{bubbles:true}));", b)
                except Exception:
                    pass
            if is_ticked():
                return True
            time.sleep(0.3)
        return False

    def present(self, xpath):
        for el in self.driver.find_elements(By.XPATH, xpath):
            try:
                if el.is_displayed():
                    return True
            except Exception:
                pass
        return False

    def visible_text(self):
        try:
            return self.driver.execute_script("return document.body.innerText") or ""
        except Exception:
            return ""

    def wait_loaded(self, timeout=30, settle=0.25):
        """Block until the page has finished loading: document ready AND no
        visible loading spinner/overlay, held stable for `settle` seconds.
        Always returns within `timeout`, so a stuck spinner can never hang the
        run. Returns True if it reached an idle state, False on timeout."""
        busy_js = r"""
        if (document.readyState === 'loading') return true;
        const sels = ['.ngx-spinner-overlay','.mat-progress-spinner','.mat-spinner',
          '.p-progress-spinner','.spinner-border','[role=progressbar]',
          '[class*=spinner]','[class*=loader]','.loading-overlay'];
        for (const s of sels){
          let els;
          try { els = document.querySelectorAll(s); } catch(e){ continue; }
          for (const el of els){
            const r = el.getBoundingClientRect();
            if (r.width === 0 || r.height === 0) continue;
            const st = window.getComputedStyle(el);
            if (st.visibility === 'hidden' || st.display === 'none' || st.opacity === '0') continue;
            return true;                 // a spinner/loader is visible
          }
        }
        return false;
        """
        end = time.time() + timeout
        idle_since = None
        while time.time() < end:
            try:
                busy = self.driver.execute_script(busy_js)
            except Exception:
                busy = False
            if busy:
                idle_since = None
            else:
                if idle_since is None:
                    idle_since = time.time()
                elif time.time() - idle_since >= settle:
                    return True
            time.sleep(0.1)
        return False

    # ---- THE LOGIN ------------------------------------------------------- #
    PAN_XPATH = ("//input[@id='panAdharUserId' or @formcontrolname='userId'"
                 " or contains(@placeholder,'PAN') or contains(@placeholder,'User ID')]")

    def _on_login_page(self, timeout=5):
        """True once the PAN/User-ID field is on screen (we're on the login
        page). After a previous row's 'Log In Again', this is already true, so
        we can skip re-navigating and just type the next PAN."""
        end = time.time() + timeout
        while time.time() < end:
            if self.present(self.PAN_XPATH):
                return True
            time.sleep(0.3)
        return False

    def login(self, pan, password, attempts=3):
        d = self.driver
        pan = str(pan).strip().upper()
        self.log(f"Logging in — PAN {pan}")

        pan_locators = [
            (By.XPATH, "//input[contains(@placeholder,'PAN') or contains(@placeholder,'AADHAAR') or contains(@placeholder,'User ID')]"),
            (By.ID, "panAdharUserId"),
            (By.XPATH, "//input[@formcontrolname='userId']"),
            (By.XPATH, "//input[@type='text' and not(@disabled)]"),
        ]
        pw_locators = [
            (By.ID, "loginPasswordField"),
            (By.XPATH, "//input[@type='password']"),
            (By.XPATH, "//input[@formcontrolname='password']"),
        ]

        # Only load the login page if we aren't already on it (the previous
        # row's 'Log In Again' click leaves us here — no reload needed). Gate the
        # wait on the URL so a fresh browser doesn't idle before the first load.
        on_login = False
        try:
            if "incometax" in (d.current_url or "").lower():
                on_login = self._on_login_page(timeout=4)
        except Exception:
            on_login = False

        if not on_login:
            d.get(self.LOGIN_URL)
            try:
                self.type(pan_locators, pan, timeout=12)
            except Exception:
                self.log("  (direct login page didn't load — homepage fallback)")
                d.get(self.PORTAL_URL)
                try:
                    self.click_text("Login", timeout=15)
                except Exception:
                    pass
                self.type(pan_locators, pan, timeout=15)
        else:
            self.type(pan_locators, pan, timeout=12)

        self.click([(By.XPATH, "//button[normalize-space()='Continue']"),
                    (By.ID, "continueBtn")])

        if self.check_box(timeout=8):
            self.log("  secure-access checkbox ticked")
        else:
            self.log("  WARNING: could not tick secure-access checkbox")

        self.type(pw_locators, password)
        status = self._submit(password, pw_locators, attempts)
        if status == "invalid":
            raise InvalidPasswordError(f"Invalid password for PAN {pan}")
        if status == "failed":
            raise LoginError(f"Login did not proceed for PAN {pan}")

        time.sleep(0.8)
        page = d.page_source.lower()
        if "otp" in page or "captcha" in page:
            self.log("  OTP/CAPTCHA detected — solve it in the browser window")
            # No console here (.pyw). Wait for the OTP screen to clear.
            self._wait_otp_cleared(timeout=180)

        self.wait_loaded(timeout=30)
        self.log("  login complete")
        return True

    def _wait_otp_cleared(self, timeout=180):
        """Poll until the OTP/CAPTCHA screen disappears (user solved it) or the
        dashboard loads. Used instead of input() because .pyw has no console."""
        end = time.time() + timeout
        while time.time() < end:
            low = self.visible_text().lower()
            if ("otp" not in low and "captcha" not in low) or "e-file" in low or "dashboard" in low:
                return True
            time.sleep(1.0)
        return False

    def _submit(self, password, pw_locators, attempts=3):
        continue_locs = [
            (By.XPATH, "//button[normalize-space()='Continue']"),
            (By.ID, "submitBtn"),
            (By.XPATH, "//button[contains(.,'Continue')]"),
        ]
        INVALID = ["invalid password", "incorrect password", "password is incorrect",
                   "invalid user id or password", "user id or password",
                   "password you entered", "wrong password"]
        REAUTH = ["request not authenticated"]

        for attempt in range(1, attempts + 1):
            if attempt > 1:
                self.log(f"  re-entering password (attempt {attempt}/{attempts})")
                self._erase_password(pw_locators)
                self.type(pw_locators, password, timeout=8)
            try:
                self.click(continue_locs, timeout=8)
            except Exception:
                self.log(f"  WARNING: Continue not clickable (attempt {attempt})")

            end = time.time() + 6
            while time.time() < end:
                self._dismiss_login_here()
                if not self.present("//input[@type='password']"):
                    return "ok"                                  # progressed — fastest exit
                low = self.visible_text().lower()
                if any(k in low for k in INVALID):
                    return "invalid"
                if any(k in low for k in REAUTH):
                    self.log("  'Request not authenticated' — will erase & retry")
                    break
                time.sleep(0.2)
        return "ok" if not self.present("//input[@type='password']") else "failed"

    def _erase_password(self, pw_locators):
        try:
            el = self.find(pw_locators, timeout=8)
        except Exception:
            return
        for action in (
            lambda: (el.click(), el.send_keys(Keys.CONTROL, "a"), el.send_keys(Keys.DELETE)),
            lambda: el.clear(),
            lambda: self.driver.execute_script(
                "arguments[0].value='';"
                "arguments[0].dispatchEvent(new Event('input',{bubbles:true}));"
                "arguments[0].dispatchEvent(new Event('change',{bubbles:true}));", el),
        ):
            try:
                action()
            except Exception:
                pass

    def _dismiss_login_here(self):
        """One-shot, non-blocking: click a 'Login here' dialog if present, else
        return immediately (never polls — so it doesn't slow the submit loop)."""
        js = r"""
        for (const el of document.querySelectorAll('button,a,span,div')){
          const t = (el.innerText || el.textContent || '').trim().toLowerCase();
          if (t === 'login here'){
            const r = el.getBoundingClientRect();
            if (r.width && r.height){ el.click(); return true; }
          }
        }
        return false;
        """
        try:
            return bool(self.driver.execute_script(js))
        except Exception:
            return False

    def quit(self):
        try:
            self.driver.quit()
        except Exception:
            pass


# =========================================================================== #
#  PART 2 — WORKFLOW  (navigate, filter by AY, read status, log out)           #
# =========================================================================== #

# Known portal statuses, most-specific first. The LATEST status is whichever
# of these sits at the TOP of the filing's status timeline.
KNOWN_STATUSES = [
    "Processed with no demand/refund",
    "Processed with no demand / refund",
    "Processed with no demand or refund",
    "Processed with refund due",
    "Processed with demand due",
    "Processed with refund",
    "Processed with demand",
    "Refund Issued",
    "Refund Failed",
    "Refund Re-issued",
    "Successfully e-Verified",
    "Submitted and pending for e-Verification",
    "Pending for e-Verification",
    "Pending for verification",
    "Under Processing",
    "ITR Processed",
    "Return uploaded",
    "Defective",
    "Invalidated",
    "Transferred to Jurisdictional Assessing Officer",
    "Rectification processed",
    "Case transferred to AO",
]


def go_to_filed_returns(portal):
    """e-File -> Income Tax Returns -> View Filed Returns."""
    portal.log("  navigating: e-File > Income Tax Returns > View Filed Returns")
    portal.wait_loaded(timeout=30)
    portal.click_css("#e-File", timeout=20)               # e-File menu (stable id)
    portal.click_menu_item("Income Tax Returns", timeout=15)   # submenu trigger
    portal.click_text("View Filed Returns", timeout=15)        # (unchanged)
    portal.wait_loaded(timeout=30)
    # Confirm the results page actually rendered.
    end = time.time() + 20
    while time.time() < end:
        low = portal.visible_text().lower()
        if "filed returns" in low or "filing type" in low or "acknowledgement" in low:
            return True
        time.sleep(0.4)
    portal.log("  WARNING: 'View Filed Returns' page did not clearly load")
    return False


def filter_by_ay(portal, ay):
    """Filter the list: click Filter (#filterbtn1) -> open the Assessment Year
    mat-select -> pick the year -> click the panel's Filter apply (#okButton)."""
    ay = normalize_ay(ay)
    portal.log(f"  filtering by Assessment Year {ay}")

    # 1) Open the Filter panel (client-side toggle — no page load to wait for).
    try:
        portal.click_css("#filterbtn1", timeout=10)
    except Exception:
        portal.log("  WARNING: could not find the Filter button")

    # 2) Open the Assessment Year dropdown and pick the year.
    if _choose_ay(portal, ay):
        portal.log(f"  Assessment Year set to {ay}")
    else:
        portal.log(f"  WARNING: could not set Assessment Year to {ay}")

    # 3) Apply (the panel's blue 'Filter' button has id 'okButton').
    try:
        portal.click_css("#okButton", timeout=8)
    except Exception:
        portal.log("  WARNING: could not click the Filter (apply) button")
    portal.wait_loaded(timeout=30)


def _choose_ay(portal, ay):
    """Open the Assessment Year mat-select (formcontrolname='ay') and click the
    matching year option, then close the (multi-select) overlay so the apply
    button is clickable."""
    d = portal.driver

    open_js = r"""
    const host = document.querySelector('mat-select[formcontrolname=ay]');
    if (!host) return false;
    const trig = host.querySelector('.mat-mdc-select-trigger') || host;
    trig.scrollIntoView({block:'center'});
    trig.click();
    return true;
    """
    pick_js = r"""
    const want = arguments[0].trim().toLowerCase().replace(/\s+/g,'');
    for (const el of document.querySelectorAll('mat-option, .mat-mdc-option, [role=option]')){
      const r = el.getBoundingClientRect();
      if (r.width === 0 || r.height === 0) continue;
      const t = (el.textContent || '').trim().toLowerCase().replace(/\s+/g,'');
      if (t === want){ el.scrollIntoView({block:'center'}); el.click(); return true; }
    }
    return false;
    """
    close_js = "const b = document.querySelector('.cdk-overlay-backdrop'); if (b) b.click();"

    for _ in range(3):
        try:
            d.execute_script(open_js)
        except Exception:
            pass
        # Poll for the option to render and click it the instant it appears.
        picked = False
        end = time.time() + 2.5
        while time.time() < end:
            try:
                if d.execute_script(pick_js, ay):
                    picked = True
                    break
            except Exception:
                pass
            time.sleep(0.1)
        if picked:
            try:
                d.execute_script(close_js)      # close the overlay (multi-select stays open)
            except Exception:
                pass
            time.sleep(0.2)
            return True
    return False


# Phrases that mean the portal returned no filing for the chosen AY.
EMPTY_HINTS = [
    "no record", "no data", "no return", "not filed", "no filing",
    "data not available", "no result", "0 filing",
]


def read_latest_status(portal, settle=6):
    """Return the LATEST status for the filtered AY, or 'ITR not filed' when no
    return is shown. The latest status is the top-most label in the timeline.
    (filter_by_ay already waited for the results to load, so we read at once.)"""
    # Report ONLY the status phrase (no dates / acknowledgement numbers): we
    # return the matched known status, not the element's raw text.
    scan_js = r"""
    const known = arguments[0];                 // original-case phrases
    let best = null;            // smallest 'top' = highest on the page = latest
    for (const el of document.querySelectorAll('span,div,p,td,li,strong,b,h3,h4')){
      const r = el.getBoundingClientRect();
      if (r.width===0 || r.height===0) continue;
      const raw = (el.innerText || el.textContent || '').trim();
      if (!raw || raw.length > 120) continue;
      const low = raw.toLowerCase();
      for (const k of known){
        if (low.includes(k.toLowerCase())){
          if (!best || r.top < best.top){ best = {text: k, top: r.top}; }
          break;
        }
      }
    }
    return best ? best.text : "";
    """
    end = time.time() + settle
    low = ""
    while time.time() < end:
        try:
            txt = portal.driver.execute_script(scan_js, KNOWN_STATUSES)
        except Exception:
            txt = ""
        if txt and txt.strip():
            return txt.strip()
        low = portal.visible_text().lower()
        if any(h in low for h in EMPTY_HINTS):
            return "ITR not filed"
        time.sleep(0.25)

    # Settled with no status text. A filing card always carries an
    # acknowledgement number, so its absence means nothing was filed.
    if "acknowledgement" in low:
        return "Status not found"
    return "ITR not filed"


def extract_fields(portal):
    """After the AY filter is applied, pull extra fields from the filing card:
    Acknowledgement No, Filing Date, and the Intimation Order date. Each field
    that is absent comes back as 'Not found'.

    Structure on the portal: every field is a pair of <mat-label> — one
    class='rightsideLabel' (the caption, e.g. 'Acknowledgement No :') followed
    by one class='fieldVal' (the value). The intimation is the exception: a
    '.hyperLink' span reading 'Download Intimation Order Dated <date>'."""
    js = r"""
    const out = {ack: "", filing: "", intimation: ""};
    const clean = (s) => (s || "").replace(/\s+/g, " ").trim();

    // 1) Caption/value pairs.
    for (const lab of document.querySelectorAll(".rightsideLabel")){
      const key = clean(lab.textContent).replace(/:/g, "").toLowerCase();
      let val = "";
      let sib = lab.nextElementSibling;
      while (sib){
        if (sib.classList && sib.classList.contains("fieldVal")){ val = clean(sib.textContent); break; }
        sib = sib.nextElementSibling;
      }
      if (!val && lab.parentElement){
        const fv = lab.parentElement.querySelector(".fieldVal");
        if (fv) val = clean(fv.textContent);
      }
      if (!val) continue;
      if (!out.ack && key.includes("acknowledgement")) out.ack = val;         // first wins
      else if (!out.filing && (key.includes("filing date") || key.includes("date of filing"))) out.filing = val;
    }

    // 2) Intimation Order date from the 'Download Intimation Order Dated ...' link.
    for (const el of document.querySelectorAll(".hyperLink, span, a")){
      const t = clean(el.textContent);
      if (!t || t.length > 90) continue;
      const m = t.match(/intimation order dated\s*(.+)/i);
      if (m){ out.intimation = m[1].trim(); break; }
    }
    return out;
    """
    # Poll briefly, but return the instant any field is found (data usually
    # appears immediately after the filter is applied).
    data = {}
    end = time.time() + 4
    while time.time() < end:
        try:
            data = portal.driver.execute_script(js) or {}
        except Exception:
            data = {}
        if data.get("ack") or data.get("filing") or data.get("intimation"):
            break
        time.sleep(0.2)

    def val(x):
        x = (str(x) if x is not None else "").strip()
        return x if x else "Not found"

    return {
        "ack": val(data.get("ack")),
        "filing_date": val(data.get("filing")),
        "intimation": val(data.get("intimation")),
    }


def logout(portal):
    """Open the profile menu (button.profileMenubtn), click 'Log Out', then
    click 'Log In Again' so we land back on the login page WITHOUT reopening the
    browser — ready for the next row."""
    portal.wait_loaded(timeout=15)

    done = False
    for _ in range(3):
        try:
            portal.click_css("button.profileMenubtn", timeout=6)   # open profile menu
        except Exception:
            pass
        time.sleep(0.6)
        try:
            portal.click_menu_item("Log Out", timeout=5)           # dropdown item
            done = True
            break
        except Exception:
            time.sleep(0.5)
    if not done:
        portal.log("  WARNING: could not click Log Out")

    # After logout the portal shows a 'Log In Again' button (registerButton).
    # click_css already polls for it, so no separate page-load wait is needed.
    try:
        portal.click_css("button.registerButton", timeout=8)
    except Exception:
        try:
            portal.click_text("Log In Again", timeout=4)
        except Exception:
            pass
    time.sleep(1.0)   # ~1s break; the next login() waits for the PAN field itself
    portal.log("  logged out")
    return done


# ---- small file/text utilities ------------------------------------------- #
_MONTHS = {m.lower(): i for i, m in enumerate(
    ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
     "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"], start=1)}


def parse_portal_date(s):
    """Parse a portal date like 'Jun 10, 2026' into a real datetime.date.
    Month names are parsed manually (not via strptime) so it works regardless
    of the machine's locale. Returns None if `s` is not a date (e.g. 'Not found')."""
    s = str(s or "").strip()
    if not s:
        return None
    m = re.match(r"([A-Za-z]{3,9})\.?\s+(\d{1,2}),?\s+(\d{4})$", s)   # Jun 10, 2026
    if m:
        mon = _MONTHS.get(m.group(1)[:3].lower())
        if mon:
            try:
                return date(int(m.group(3)), mon, int(m.group(2)))
            except Exception:
                return None
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%b-%Y", "%d %b %Y"):  # numeric fallbacks
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            continue
    return None


def fmt_date_display(s):
    """Portal date -> 'DD-MM-YYYY' for the log; unchanged if not a date."""
    d = parse_portal_date(s)
    return d.strftime("%d-%m-%Y") if d else s


def normalize_ay(ay):
    """'AY 2026-27' / '2026-27' / '2026-2027' -> '2026-27'."""
    s = str(ay).strip()
    s = re.sub(r"(?i)^a\.?y\.?\s*", "", s).strip()
    m = re.search(r"(20\d{2})\s*[-/]\s*(\d{2,4})", s)
    if m:
        start, end = m.group(1), m.group(2)
        if len(end) == 4:
            end = end[-2:]
        return f"{start}-{end}"
    return s


# =========================================================================== #
#  VERIFIED (Yes/No) — with a small self-training knowledge base                #
# =========================================================================== #
# Default rule: a return is "verified" if its status shows it was processed or
# e-verified, OR an intimation order exists. The tool also LEARNS: every new
# status it meets is recorded (with its verdict) in a JSON file you can review
# and correct, so future runs reuse your decisions.

VERIFIED_KEYWORDS = ("processed", "e-verified", "everified", "verified")

# Statuses that carry no verification info — always 'No', never learned.
_NON_STATUS = ("invalid password", "login failed", "itr not filed",
               "status not found", "error")


def verified_rules_path():
    d = os.path.join(os.path.expanduser("~"), ".itr_status_checker")
    try:
        os.makedirs(d, exist_ok=True)
    except Exception:
        pass
    return os.path.join(d, "verified_rules.json")


def load_verified_rules():
    """Return {normalized status -> 'Yes'/'No'} learned so far (or {})."""
    try:
        with open(verified_rules_path(), "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, dict):
            return {_norm_status(k): ("Yes" if str(v).strip().lower().startswith("y") else "No")
                    for k, v in data.items() if str(k).strip()}
    except Exception:
        pass
    return {}


def save_verified_rules(rules):
    try:
        with open(verified_rules_path(), "w", encoding="utf-8") as f:
            json.dump(rules, f, indent=2, ensure_ascii=False, sort_keys=True)
        return True
    except Exception:
        return False


def _norm_status(s):
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


def classify_verified(status, intimation, rules=None):
    """Return 'Yes'/'No'. An intimation order always means verified; otherwise
    consult the learned rules, then the default keyword heuristic."""
    rules = rules or {}
    s = _norm_status(status)
    intim = _norm_status(intimation)
    if intim and intim != "not found":
        return "Yes"
    if not s or any(s.startswith(k) for k in _NON_STATUS):
        return "No"
    if s in rules:                       # user/learned decision wins
        return rules[s]
    if any(k in s for k in VERIFIED_KEYWORDS):
        return "Yes"
    return "No"


def learnable_status(status):
    """True if this status is a real portal status worth remembering."""
    s = _norm_status(status)
    return bool(s) and not any(s.startswith(k) for k in _NON_STATUS)


# =========================================================================== #
#  PART 3 — EXCEL  (template export + read rows + export updated copy)         #
# =========================================================================== #
HEADERS = ["Name", "PAN", "Password", "AY", "Status",
           "Acknowledgement No", "Filing Date", "Intimation", "Verified"]

# Internal field key -> column header, for the values the tool fills in itself.
OUTPUT_COLS = [
    ("status", "Status"),
    ("ack", "Acknowledgement No"),
    ("filing_date", "Filing Date"),
    ("intimation", "Intimation"),
    ("verified", "Verified"),
]


def export_template(path):
    """Write a blank, formatted template workbook to `path`."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Returns"

    head_fill = PatternFill("solid", fgColor="1F4E78")
    head_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # A sample row (greyed) to show the expected format.
    sample = ["John Doe", "ABCDE1234F", "YourPassword", "2026-27", "", "", "", "", ""]
    for c, v in enumerate(sample, start=1):
        cell = ws.cell(row=2, column=c, value=v)
        cell.font = Font(italic=True, color="9C9C9C")
        cell.border = border

    widths = [24, 16, 18, 12, 28, 22, 16, 24, 10]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"

    # Instructions sheet.
    info = wb.create_sheet("Instructions")
    notes = [
        "HOW TO USE THIS TEMPLATE",
        "",
        "1. Fill ONE row per return you want to check.",
        "2. Columns:",
        "      Name      - any label to identify the taxpayer (for your reference only)",
        "      PAN       - 10-character PAN, e.g. ABCDE1234F",
        "      Password  - the e-Filing portal login password for that PAN",
        "      AY        - Assessment Year to filter by, e.g. 2026-27",
        "      Status              - LEAVE BLANK. Filled in automatically.",
        "      Acknowledgement No  - LEAVE BLANK. Filled in automatically.",
        "      Filing Date         - LEAVE BLANK. Filled in automatically.",
        "      Intimation          - LEAVE BLANK. Filled in automatically.",
        "      Verified            - LEAVE BLANK. Yes/No, filled automatically.",
        "",
        "3. Delete the grey sample row before running.",
        "4. Save the file, then upload it in the tool and click Run.",
        "",
        "The tool logs in for each row, applies the Assessment Year filter, then",
        "reads the latest status plus the Acknowledgement No, Filing Date and",
        "Intimation Order date. Anything not present is recorded as 'Not found'.",
        "It also marks Verified = Yes/No (Yes when the return is processed or",
        "e-verified, or an intimation exists).",
        "",
        "When the run finishes, THIS SAME FILE is updated in place with the",
        "results — no separate download needed. (Keep it closed while running.)",
    ]
    for r, line in enumerate(notes, start=1):
        cell = info.cell(row=r, column=1, value=line)
        if r == 1:
            cell.font = Font(bold=True, size=12, color="1F4E78")
    info.column_dimensions["A"].width = 90

    wb.save(path)


def read_rows(path):
    """Return (workbook, worksheet, col_map, rows). rows is a list of dicts
    with keys name/pan/password/ay and an 'excel_row' index."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Map headers (case-insensitive) on row 1.
    col_map = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=c).value
        if h:
            col_map[str(h).strip().lower()] = c
    for needed in ("pan", "password", "ay"):
        if needed not in col_map:
            raise ValueError(f"Column '{needed.upper()}' not found in the sheet header.")
    # Make sure every output column exists (append any that are missing, so old
    # templates without the new columns still work).
    for _key, header in OUTPUT_COLS:
        lk = header.strip().lower()
        if lk not in col_map:
            col_map[lk] = ws.max_column + 1
            ws.cell(row=1, column=col_map[lk], value=header)

    rows = []
    for r in range(2, ws.max_row + 1):
        pan = ws.cell(row=r, column=col_map["pan"]).value
        pw = ws.cell(row=r, column=col_map["password"]).value
        if not pan or not pw:
            continue
        pan_s = str(pan).strip()
        # skip the grey sample row from the template
        if pan_s.upper() == "ABCDE1234F":
            continue
        rows.append({
            "excel_row": r,
            "name": str(ws.cell(row=r, column=col_map.get("name", 0)).value or "").strip()
                    if col_map.get("name") else "",
            "pan": pan_s,
            "password": str(pw),
            "ay": str(ws.cell(row=r, column=col_map["ay"]).value or "").strip(),
        })
    return wb, ws, col_map, rows


def save_updated_excel(src_path, results, dst_path):
    """Load the uploaded workbook, write the collected values (Status,
    Acknowledgement No, Filing Date, Intimation) into their columns, and save a
    NEW copy to dst_path. The original file is left untouched. `results` maps
    excel_row -> dict of field values."""
    wb, ws, col_map, _ = read_rows(src_path)
    for excel_row, data in results.items():
        if not isinstance(data, dict):
            data = {"status": data}                          # tolerate a bare status
        for key, header in OUTPUT_COLS:
            idx = col_map.get(header.strip().lower())
            if idx:
                ws.cell(row=excel_row, column=idx, value=data.get(key, "Not found"))
    wb.save(dst_path)


# =========================================================================== #
#  PART 4 — THE BATCH WORKER (runs on a background thread)                      #
# =========================================================================== #
def run_batch(excel_path, headless, log, stop_event, on_result):
    if not _OPENPYXL_OK:
        log("ERROR: openpyxl is not installed. Run:  pip install openpyxl")
        return

    try:
        _, _, _, rows = read_rows(excel_path)
    except Exception as e:
        log(f"ERROR reading Excel: {e}")
        return

    if not rows:
        log("No data rows found. Fill the template (and remove the grey sample row).")
        return

    log(f"Loaded {len(rows)} row(s).")
    log("=" * 60)

    rules = load_verified_rules()        # learned Yes/No decisions so far
    rules_changed = False
    collected = {}                       # excel_row -> data (for writing back)
    lock_warned = [False]                # so we warn about a locked file only once

    # One browser for the whole batch — we log out and click 'Log In Again'
    # between rows instead of reopening Edge each time (much faster).
    portal = None
    try:
        portal = IncomeTaxPortal(headless=headless, log=log)

        for i, row in enumerate(rows, start=1):
            if stop_event.is_set():
                log("Stopped by user.")
                break

            name, pan, pw, ay = row["name"], row["pan"], row["password"], row["ay"]
            log(f"[{i}/{len(rows)}] {name or pan}  (PAN {pan}, AY {ay or '—'})")

            data = {"status": "Error", "ack": "Not found", "filing_date": "Not found",
                    "intimation": "Not found", "verified": "No"}
            try:
                portal.login(pan, pw)

                go_to_filed_returns(portal)
                if ay:
                    filter_by_ay(portal, ay)

                data["status"] = read_latest_status(portal)
                log(f"  STATUS: {data['status']}")

                # Only pull the extra fields when a filing actually exists.
                if data["status"] not in ("ITR not filed", "Status not found"):
                    data.update(extract_fields(portal))
                log(f"  Ack: {data['ack']}  |  Filed: {data['filing_date']}  |  "
                    f"Intimation: {data['intimation']}")

                logout(portal)          # ends on the login page ('Log In Again')

            except InvalidPasswordError:
                data["status"] = "Invalid password"
                log("  STATUS: Invalid password — skipping")
                _recover_to_login(portal)
            except LoginError as e:
                data["status"] = "Login failed"
                log(f"  STATUS: Login failed ({e})")
                _recover_to_login(portal)
            except Exception as e:
                data["status"] = f"Error: {e}"
                log(f"  ERROR: {e}")
                _recover_to_login(portal)
            finally:
                # Verified Yes/No + self-training on any newly-seen status.
                data["verified"] = classify_verified(data["status"], data["intimation"], rules)
                if learnable_status(data["status"]) and _norm_status(data["status"]) not in rules:
                    rules[_norm_status(data["status"])] = data["verified"]
                    rules_changed = True
                log(f"  Verified: {data['verified']}")

                collected[row["excel_row"]] = data
                on_result(row["excel_row"], data)
                _write_back(excel_path, collected, log, lock_warned)   # update template as we go
            log("-" * 60)
    finally:
        if portal:
            portal.quit()

    if rules_changed:
        save_verified_rules(rules)

    # Final write-back into the uploaded template (report the outcome clearly).
    if collected:
        if _write_back(excel_path, collected, log, [False], final=True):
            log(f"Your template was updated in place:\n  {excel_path}")
        else:
            log("Could NOT update your template (it may be open in Excel). "
                "Close it, then click 'Save to my template'.")

    log("=" * 60)
    log("Done.")


def _write_back(excel_path, collected, log, lock_warned, final=False):
    """Write the collected results straight back into the uploaded template.
    Returns True on success. On a locked file (open in Excel) it warns once
    (unless final) and returns False — the run keeps going regardless."""
    try:
        save_updated_excel(excel_path, collected, excel_path)
        return True
    except PermissionError:
        if final or not lock_warned[0]:
            log("  NOTE: template is open in Excel — results kept in memory; "
                "close it to let the tool save.")
            lock_warned[0] = True
        return False
    except Exception as e:
        if final or not lock_warned[0]:
            log(f"  NOTE: couldn't update the template ({e}).")
            lock_warned[0] = True
        return False


def _recover_to_login(portal):
    """After a failed row, return the SAME browser to a clean login page for the
    next row. Try a normal logout first; if that doesn't land us on the login
    page, clear the session and reload it (never reopens the browser)."""
    try:
        logout(portal)
    except Exception:
        pass
    try:
        if portal._on_login_page(timeout=3):
            return
    except Exception:
        pass
    try:
        portal.driver.delete_all_cookies()
    except Exception:
        pass
    try:
        portal.driver.get(portal.LOGIN_URL)
    except Exception:
        pass


# =========================================================================== #
#  PART 5 — TKINTER GUI                                                         #
# =========================================================================== #
class App:
    def __init__(self, root):
        self.root = root
        root.title("Income-Tax Returns — Bulk Status Checker")
        root.geometry("760x560")
        root.minsize(680, 480)

        self.excel_path = tk.StringVar(value="")
        self.headless = tk.BooleanVar(value=False)
        self.results = {}            # excel_row -> status

        self.log_queue = queue.Queue()
        self.stop_event = threading.Event()
        self.worker = None

        self._build_ui()
        self._poll_log()

        if not _OPENPYXL_OK:
            self._log("WARNING: 'openpyxl' is not installed — the Excel features "
                      "will not work.\n         Install it with:  pip install openpyxl")

    # ---- layout ----------------------------------------------------------- #
    def _build_ui(self):
        pad = dict(padx=10, pady=6)

        head = ttk.Label(self.root, text="Bulk verification-status checker for filed Income-Tax returns",
                         font=("Segoe UI", 12, "bold"))
        head.pack(anchor="w", padx=12, pady=(12, 2))
        ttk.Label(self.root,
                  text="Flow:  download template  ›  fill it  ›  upload  ›  Run  (your file is updated in place)",
                  foreground="#555").pack(anchor="w", padx=12, pady=(0, 8))

        # Step 1 — template
        f1 = ttk.LabelFrame(self.root, text="Step 1 — Get the Excel template")
        f1.pack(fill="x", **pad)
        ttk.Button(f1, text="Download Excel Template…",
                   command=self.on_template).pack(side="left", padx=8, pady=8)
        ttk.Label(f1, text="Columns you fill: Name · PAN · Password · AY   (Status … Verified are auto-filled)",
                  foreground="#555").pack(side="left", padx=6)

        # Step 2 — upload
        f2 = ttk.LabelFrame(self.root, text="Step 2 — Upload your filled template")
        f2.pack(fill="x", **pad)
        ttk.Button(f2, text="Upload Excel…", command=self.on_upload).pack(side="left", padx=8, pady=8)
        ttk.Label(f2, textvariable=self.excel_path, foreground="#1F4E78").pack(side="left", padx=6)

        # Step 3 — run
        f4 = ttk.LabelFrame(self.root, text="Step 3 — Run")
        f4.pack(fill="x", **pad)
        self.run_btn = ttk.Button(f4, text="▶  Run", command=self.on_run)
        self.run_btn.pack(side="left", padx=8, pady=8)
        self.stop_btn = ttk.Button(f4, text="■  Stop", command=self.on_stop, state="disabled")
        self.stop_btn.pack(side="left", padx=4, pady=8)
        ttk.Checkbutton(f4, text="Headless (hide browser — not recommended; OTP/CAPTCHA need the window)",
                        variable=self.headless).pack(side="left", padx=12)

        # Step 4 — results (written back into your uploaded file) + training
        f5 = ttk.LabelFrame(self.root, text="Step 4 — Results & training")
        f5.pack(fill="x", **pad)
        self.save_btn = ttk.Button(f5, text="Save to my template",
                                   command=self.on_save_to_template, state="disabled")
        self.save_btn.pack(side="left", padx=8, pady=8)
        ttk.Button(f5, text="Verified rules (train)…",
                   command=self.on_train_verified).pack(side="left", padx=4, pady=8)
        ttk.Label(f5, text="Results are written straight into your uploaded file. "
                           "'Save' re-writes it if it was open in Excel.",
                  foreground="#555").pack(side="left", padx=6)

        # Log
        flog = ttk.LabelFrame(self.root, text="Progress log")
        flog.pack(fill="both", expand=True, **pad)
        self.log_box = scrolledtext.ScrolledText(flog, height=12, wrap="word",
                                                 font=("Consolas", 9))
        self.log_box.pack(fill="both", expand=True, padx=6, pady=6)
        self.log_box.configure(state="disabled")

    # ---- button handlers -------------------------------------------------- #
    def on_template(self):
        if not _OPENPYXL_OK:
            messagebox.showerror("Missing dependency",
                                 "openpyxl is not installed.\n\nRun:  pip install openpyxl")
            return
        path = filedialog.asksaveasfilename(
            title="Save template as…",
            defaultextension=".xlsx",
            initialfile="ITR_status_template.xlsx",
            filetypes=[("Excel workbook", "*.xlsx")])
        if not path:
            return
        try:
            export_template(path)
            self._log(f"Template saved: {path}")
            messagebox.showinfo("Template saved",
                                f"Template created:\n{path}\n\nFill it, then upload it in Step 2.")
        except Exception as e:
            messagebox.showerror("Could not save template", str(e))

    def on_upload(self):
        path = filedialog.askopenfilename(
            title="Select your filled template",
            filetypes=[("Excel workbook", "*.xlsx *.xlsm"), ("All files", "*.*")])
        if not path:
            return
        self.excel_path.set(path)
        self._log(f"Loaded Excel: {path}")
        # quick preview of row count
        if _OPENPYXL_OK:
            try:
                _, _, _, rows = read_rows(path)
                self._log(f"  {len(rows)} data row(s) detected.")
            except Exception as e:
                self._log(f"  WARNING: {e}")

    def on_save_to_template(self):
        """Write the collected results back into the uploaded file (in place).
        Used as a retry if the file was open in Excel during the run."""
        if not self.results:
            messagebox.showinfo("Nothing yet",
                                "Run the batch first — there are no results to save.")
            return
        path = self.excel_path.get()
        try:
            save_updated_excel(path, dict(self.results), path)
            self._log(f"Saved results into your template: {path}")
            messagebox.showinfo("Saved", f"Your template was updated:\n{path}")
        except PermissionError:
            messagebox.showerror(
                "File is open",
                "Your template is open in Excel. Close it and try again.")
        except Exception as e:
            messagebox.showerror("Could not save", str(e))

    def on_train_verified(self):
        """Open a small dialog to review and edit the learned Verified (Yes/No)
        rules — the tool's self-training knowledge base."""
        rules = load_verified_rules()

        win = tk.Toplevel(self.root)
        win.title("Verified rules — train Yes/No")
        win.geometry("560x460")
        win.transient(self.root)

        ttk.Label(win, text="Teach the tool which statuses count as Verified. "
                            "Intimations always count as Yes.",
                  wraplength=520, foreground="#333").pack(anchor="w", padx=12, pady=(12, 4))

        listwrap = ttk.Frame(win)
        listwrap.pack(fill="both", expand=True, padx=12, pady=4)
        lb = tk.Listbox(listwrap, font=("Consolas", 10), activestyle="dotbox")
        sb = ttk.Scrollbar(listwrap, orient="vertical", command=lb.yview)
        lb.configure(yscrollcommand=sb.set)
        lb.pack(side="left", fill="both", expand=True)
        sb.pack(side="left", fill="y")

        order = []            # parallel list of status keys, matching listbox rows

        def refresh(select=None):
            lb.delete(0, "end")
            order.clear()
            for k in sorted(rules):
                order.append(k)
                lb.insert("end", f"{rules[k]:>3}   |   {k}")
            if select in order:
                idx = order.index(select)
                lb.selection_set(idx)
                lb.see(idx)

        # editor row
        editor = ttk.Frame(win)
        editor.pack(fill="x", padx=12, pady=(4, 2))
        ttk.Label(editor, text="Status:").pack(side="left")
        status_var = tk.StringVar()
        ttk.Entry(editor, textvariable=status_var, width=40).pack(side="left", padx=6)
        verdict_var = tk.StringVar(value="Yes")
        ttk.Radiobutton(editor, text="Yes", variable=verdict_var, value="Yes").pack(side="left")
        ttk.Radiobutton(editor, text="No", variable=verdict_var, value="No").pack(side="left")

        def on_select(_evt=None):
            sel = lb.curselection()
            if sel:
                k = order[sel[0]]
                status_var.set(k)
                verdict_var.set(rules.get(k, "Yes"))
        lb.bind("<<ListboxSelect>>", on_select)

        def add_update():
            k = _norm_status(status_var.get())
            if not k:
                return
            rules[k] = verdict_var.get()
            refresh(select=k)

        def remove():
            k = _norm_status(status_var.get())
            if k in rules:
                del rules[k]
                status_var.set("")
                refresh()

        btns = ttk.Frame(win)
        btns.pack(fill="x", padx=12, pady=(2, 6))
        ttk.Button(btns, text="Add / Update", command=add_update).pack(side="left")
        ttk.Button(btns, text="Remove", command=remove).pack(side="left", padx=6)

        def save_close():
            if save_verified_rules(rules):
                self._log(f"Saved {len(rules)} verified rule(s).")
            else:
                messagebox.showerror("Could not save", "Failed to save the rules file.")
            win.destroy()

        foot = ttk.Frame(win)
        foot.pack(fill="x", padx=12, pady=(0, 12))
        ttk.Button(foot, text="Save & Close", command=save_close).pack(side="right")
        ttk.Button(foot, text="Cancel", command=win.destroy).pack(side="right", padx=6)

        refresh()

    def on_run(self):
        if self.worker and self.worker.is_alive():
            return
        if not self.excel_path.get():
            messagebox.showwarning("Missing", "Upload your filled Excel template first (Step 2).")
            return
        if not os.path.isfile(self.excel_path.get()):
            messagebox.showerror("Not found", "The Excel file no longer exists.")
            return

        if not messagebox.askyesno(
                "Start batch?",
                "An Edge window will open and log in for each row in turn.\n\n"
                "Keep the browser visible so you can solve any OTP/CAPTCHA.\n"
                "Your results are written back into this same file — please keep "
                "it CLOSED in Excel while the run is in progress.\n\n"
                "Start now?"):
            return

        self.results = {}
        self.stop_event.clear()
        self.run_btn.configure(state="disabled")
        self.stop_btn.configure(state="normal")
        self.save_btn.configure(state="disabled")
        self._log("\n>>> Starting batch…\n")

        self.worker = threading.Thread(
            target=self._worker_wrap,
            args=(self.excel_path.get(), self.headless.get()),
            daemon=True)
        self.worker.start()

    def on_stop(self):
        self.stop_event.set()
        self._log("Stop requested — will halt after the current row…")
        self.stop_btn.configure(state="disabled")

    def _worker_wrap(self, excel_path, headless):
        try:
            run_batch(excel_path, headless,
                      log=lambda m: self.log_queue.put(m),
                      stop_event=self.stop_event,
                      on_result=self._store_result)
        except Exception as e:
            self.log_queue.put(f"FATAL: {e}")
        finally:
            self.log_queue.put("__DONE__")

    def _store_result(self, excel_row, data):
        self.results[excel_row] = data

    # ---- thread-safe logging --------------------------------------------- #
    def _log(self, msg):
        self.log_box.configure(state="normal")
        self.log_box.insert("end", str(msg) + "\n")
        self.log_box.see("end")
        self.log_box.configure(state="disabled")

    def _poll_log(self):
        try:
            while True:
                msg = self.log_queue.get_nowait()
                if msg == "__DONE__":
                    self.run_btn.configure(state="normal")
                    self.stop_btn.configure(state="disabled")
                    if self.results:
                        self.save_btn.configure(state="normal")
                else:
                    self._log(msg)
        except queue.Empty:
            pass
        self.root.after(120, self._poll_log)


def main():
    root = tk.Tk()
    try:
        ttk.Style().theme_use("vista")   # nicer on Windows; ignored elsewhere
    except Exception:
        pass
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
